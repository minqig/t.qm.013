"""
T.QM.013 检验指导书自动生成工具 (自包含版)
只需一个 app.py + T.QM.013 模板文件即可运行
"""

import os
import uuid
from datetime import datetime

from flask import Flask, request, jsonify, send_file, session
from werkzeug.utils import secure_filename
import openpyxl

# ==================== 配置 ====================
app = Flask(__name__)
app.secret_key = 'qm013-secret-key-change-in-production'

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
# ★ 改成你实际的模板文件名
TEMPLATE_FILE = 'T.QM.013.xlsm'
ALLOWED_EXTENSIONS = {'xlsx', 'xlsm', 'xls'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# ==================== T.QM.013 模板单元格映射 ====================
TQM013_CELL_MAP = {
    'project':        (5, 6),
    'sub_project':    (5, 9),
    'language':       (5, 13),
    'oem':            (7, 2),
    'part_no_oem':    (7, 6),
    'dir':            (7, 9),
    'dmba':           (7, 13),
    'page':           (7, 16),
    'part_name':      (8, 2),
    'model_type':     (8, 6),
    'bat_material':   (8, 9),
    'product_group':  (8, 13),
    'instruction':    (8, 16),
    'single_part_desc': (9, 2),
    'single_part_no': (9, 10),
    'release_date':   (11, 2),
    'doc_no':         (11, 7),
    'workstation':    (11, 10),
    'name_dept':      (12, 2),
    'plan_date':      (12, 7),
    'content_start_row': 18,
    'content_cols': {
        'content_c': 3, 'content_d': 4,
        'description_1': 5, 'description_2': 6, 'description_3': 7,
        'description_4': 8, 'description_5': 9, 'description_6': 10,
        'test_level_eq': 12, 'responsible': 14,
    },
    'content_end_row': 48,
}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def parse_control_plan(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    header_row = 1
    workstation_col_idx = None

    for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row or 100)):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip().lower()
                if any(kw in val for kw in ['工位', 'op', '工作站', 'station', 'arbeitsplatz']):
                    header_row = cell.row
                    workstation_col_idx = cell.column
                    break
        if workstation_col_idx:
            break

    headers = {}
    for col_idx in range(1, (ws.max_column or 1) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        headers[col_idx] = str(cell.value).strip() if cell.value else ''

    data_start_row = header_row + 1
    all_data = []
    workstations = []

    for row_idx in range(data_start_row, (ws.max_row or data_start_row) + 1):
        row_data = {}
        has_data = False
        for col_idx in range(1, (ws.max_column or 1) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data[col_idx] = cell.value
            if cell.value is not None:
                has_data = True
        if has_data:
            all_data.append(row_data)
            if workstation_col_idx and row_data.get(workstation_col_idx):
                ws_name = str(row_data[workstation_col_idx]).strip()
                if ws_name and ws_name not in workstations:
                    workstations.append(ws_name)

    wb.close()
    return {
        'headers': headers,
        'workstations': workstations,
        'workstation_col': workstation_col_idx,
        'data': all_data,
        'header_row': header_row,
        'data_start_row': data_start_row,
        'filepath': filepath,
    }


def smart_match_columns(cp_headers):
    rules = {
        'content_c':      ['内容', 'content', 'inhalt', '工序', 'process', 'vorgang', 'nr'],
        'content_d':      ['d', '分类', 'class', 'klasse'],
        'description_1':  ['描述', 'description', 'beschreibung', '要求', 'requirement', '说明', '特征'],
        'description_2':  ['规格', 'spec', 'spezifikation', '值', 'value', '标准值'],
        'description_3':  ['方法', 'method', 'methode', '公差', 'tolerance'],
        'description_4':  ['备注', 'remark', 'bemerkung', '注释', 'note'],
        'description_5':  ['参考', 'reference', 'referenz', '文件', 'document'],
        'description_6':  ['标准', 'standard', 'norm', '规范'],
        'test_level_eq':  ['试验', '设备', 'test', 'equipment', 'prüf', 'messmittel', '检测', '测量', '量具'],
        'responsible':    ['负责人', 'responsible', 'verantwortlich', '责任', '检验人', '执行人', '部门'],
    }
    mapping = {}
    used = set()
    for tqm_col, keywords in rules.items():
        for cp_idx, cp_header in cp_headers.items():
            if cp_idx in used:
                continue
            cp_lower = cp_header.lower()
            if any(kw in cp_lower for kw in keywords):
                mapping[tqm_col] = cp_idx
                used.add(cp_idx)
                break
    return mapping


def fill_template(control_plan_data, selected_ws, column_mapping, output_path):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError(
            f"T.QM.013 模板文件未找到: {TEMPLATE_FILE}\n"
            f"请将模板文件放在: {os.path.abspath(TEMPLATE_FILE)}"
        )

    wb = openpyxl.load_workbook(TEMPLATE_FILE)
    ws = wb.active
    cell_map = TQM013_CELL_MAP
    ws_col = control_plan_data['workstation_col']

    # 填充默认值
    ws.cell(row=cell_map['language'][0], column=cell_map['language'][1], value='中文')
    ws.cell(row=cell_map['dmba'][0], column=cell_map['dmba'][1], value='A')
    ws.cell(row=cell_map['instruction'][0], column=cell_map['instruction'][1], value='ON')
    ws.cell(row=cell_map['workstation'][0], column=cell_map['workstation'][1], value=selected_ws)

    # 筛选匹配数据
    matched_rows = []
    for row_data in control_plan_data['data']:
        if ws_col and row_data.get(ws_col) is not None:
            cell_val = str(row_data[ws_col]).strip()
            if cell_val == str(selected_ws).strip():
                matched_rows.append(row_data)
            elif str(selected_ws).strip().lower() in cell_val.lower():
                matched_rows.append(row_data)

    # 填充内容
    content_cols = cell_map['content_cols']
    current_row = cell_map['content_start_row']
    for row_data in matched_rows:
        if current_row > cell_map['content_end_row']:
            break
        for tqm_col, cp_col_idx in column_mapping.items():
            if tqm_col in content_cols:
                tqm_col_idx = content_cols[tqm_col]
                value = row_data.get(cp_col_idx)
                if value is not None:
                    ws.cell(row=current_row, column=tqm_col_idx, value=value)
        current_row += 1

    wb.save(output_path)
    wb.close()
    return output_path


# ==================== 路由 ====================

@app.route('/')
def index():
    return HTML_PAGE  # 内嵌 HTML


@app.route('/api/upload_control_plan', methods=['POST'])
def upload_control_plan():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未选择文件'})
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '文件名为空'})
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': '不支持的文件格式'})

    session_id = str(uuid.uuid4())
    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
    file.save(filepath)

    try:
        cp_data = parse_control_plan(filepath)
        session['cp_filepath'] = filepath
        session['cp_session_id'] = session_id
        return jsonify({
            'success': True,
            'workstations': cp_data['workstations'],
            'headers': {str(k): v for k, v in cp_data['headers'].items()},
            'total_rows': len(cp_data['data']),
            'workstation_column': cp_data['workstation_col'],
            'session_id': session_id,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'解析失败: {str(e)}'})


@app.route('/api/get_column_mapping', methods=['POST'])
def get_column_mapping():
    data = request.get_json()
    session_id = data.get('session_id')
    cp_filepath = session.get('cp_filepath')
    if not cp_filepath or not os.path.exists(cp_filepath):
        return jsonify({'success': False, 'error': '会话已过期'})

    cp_data = parse_control_plan(cp_filepath)
    mapping = smart_match_columns(cp_data['headers'])

    tqm_labels = {
        'content_c': '内容 C', 'content_d': '内容 D',
        'description_1': '描述 1', 'description_2': '描述 2',
        'description_3': '描述 3', 'description_4': '描述 4',
        'description_5': '描述 5', 'description_6': '描述 6',
        'test_level_eq': '试验等级/设备', 'responsible': '负责人',
    }
    mapping_display = {}
    for tqm_col, cp_idx in mapping.items():
        mapping_display[tqm_col] = {
            'tqm_label': tqm_labels.get(tqm_col, tqm_col),
            'cp_column': str(cp_idx),
            'cp_header': cp_data['headers'].get(cp_idx, ''),
        }
    unmapped = [{'tqm_col': k, 'tqm_label': v} for k, v in tqm_labels.items() if k not in mapping]

    return jsonify({
        'success': True,
        'mapping': {k: v for k, v in mapping.items()},
        'mapping_display': mapping_display,
        'unmapped': unmapped,
        'all_cp_headers': {str(k): v for k, v in cp_data['headers'].items()},
    })


@app.route('/api/generate', methods=['POST'])
def generate():
    data = request.get_json()
    session_id = data.get('session_id')
    workstation = data.get('workstation')
    custom_mapping = data.get('mapping', {})

    if not session_id or not workstation:
        return jsonify({'success': False, 'error': '缺少必要参数'})

    cp_filepath = session.get('cp_filepath')
    if not cp_filepath or not os.path.exists(cp_filepath):
        return jsonify({'success': False, 'error': '会话已过期'})

    try:
        cp_data = parse_control_plan(cp_filepath)
        column_mapping = {}
        if custom_mapping:
            column_mapping = {k: int(v) for k, v in custom_mapping.items() if v}
        else:
            column_mapping = smart_match_columns(cp_data['headers'])

        safe_ws = secure_filename(str(workstation))
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"T.QM.013_{safe_ws}_{timestamp}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)

        fill_template(cp_data, workstation, column_mapping, output_path)

        session['output_path'] = output_path
        session['output_filename'] = output_filename

        return jsonify({
            'success': True,
            'filename': output_filename,
            'download_url': f'/api/download/{output_filename}',
        })
    except FileNotFoundError as e:
        return jsonify({'success': False, 'error': str(e)})
    except Exception as e:
        return jsonify({'success': False, 'error': f'生成失败: {str(e)}'})


@app.route('/api/download/<filename>')
def download(filename):
    safe_name = secure_filename(filename)
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], safe_name)
    if not os.path.exists(filepath):
        return jsonify({'success': False, 'error': '文件不存在'}), 404
    return send_file(filepath, as_attachment=True, download_name=safe_name)


@app.route('/api/cleanup', methods=['POST'])
def cleanup():
    sid = session.get('cp_session_id')
    if sid:
        for folder in [UPLOAD_FOLDER, OUTPUT_FOLDER]:
            for f in os.listdir(folder):
                if f.startswith(sid):
                    try:
                        os.remove(os.path.join(folder, f))
                    except OSError:
                        pass
    return jsonify({'success': True})


# ==================== 内嵌 HTML 页面 ====================
HTML_PAGE = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>T.QM.013 检验指导书生成器</title>
<style>
:root {
    --primary: #2563eb; --primary-hover: #1d4ed8;
    --bg: #f8fafc; --card-bg: #ffffff; --border: #e2e8f0;
    --text: #1e293b; --text-secondary: #64748b;
    --success: #16a34a; --warning: #ea580c; --danger: #dc2626;
    --radius: 8px; --shadow: 0 1px 3px rgba(0,0,0,0.1);
}
* { margin: 0; padding: 0; box-sizing: border-box; }
body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'PingFang SC', 'Microsoft YaHei', sans-serif;
    background: var(--bg); color: var(--text); line-height: 1.6; min-height: 100vh;
}
.header {
    background: var(--primary); color: white; padding: 16px 32px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.15);
}
.header h1 { font-size: 1.4rem; font-weight: 600; }
.header .subtitle { font-size: 0.85rem; opacity: 0.85; margin-top: 2px; }
.container { max-width: 1100px; margin: 0 auto; padding: 24px 16px; }

.steps {
    display: flex; gap: 0; margin-bottom: 32px;
    background: var(--card-bg); border-radius: var(--radius); box-shadow: var(--shadow); overflow: hidden;
}
.step {
    flex: 1; padding: 14px 20px; text-align: center; font-size: 0.9rem;
    color: var(--text-secondary); background: var(--card-bg); transition: all 0.3s;
}
.step.active { color: white; background: var(--primary); font-weight: 600; }
.step.done { color: var(--success); background: #f0fdf4; }
.step .step-num {
    display: inline-block; width: 24px; height: 24px; line-height: 24px;
    border-radius: 50%; border: 2px solid currentColor; margin-right: 6px;
    font-size: 0.8rem; font-weight: 700;
}
.step.active .step-num { background: white; color: var(--primary); border-color: white; }
.step.done .step-num { background: var(--success); color: white; border-color: var(--success); }

.card {
    background: var(--card-bg); border-radius: var(--radius);
    box-shadow: var(--shadow); padding: 24px; margin-bottom: 20px;
}
.card h2 {
    font-size: 1.1rem; margin-bottom: 16px; padding-bottom: 10px;
    border-bottom: 2px solid var(--border); display: flex; align-items: center; gap: 8px;
}

.upload-zone {
    border: 2px dashed var(--border); border-radius: var(--radius);
    padding: 40px; text-align: center; cursor: pointer; transition: all 0.3s; background: #fafbfc;
}
.upload-zone:hover, .upload-zone.drag-over { border-color: var(--primary); background: #eff6ff; }
.upload-zone .upload-icon { font-size: 3rem; margin-bottom: 12px; color: var(--text-secondary); }
.upload-zone p { color: var(--text-secondary); font-size: 0.95rem; }
.upload-zone .browse-link { color: var(--primary); font-weight: 600; cursor: pointer; }
.upload-zone input[type="file"] { display: none; }

.file-info {
    display: none; margin-top: 12px; padding: 10px 16px;
    background: #f0fdf4; border-radius: 6px; color: var(--success); font-size: 0.9rem;
    align-items: center; gap: 8px;
}
.file-info.show { display: flex; }

.btn {
    display: inline-flex; align-items: center; gap: 6px;
    padding: 10px 24px; border: none; border-radius: 6px;
    font-size: 0.95rem; font-weight: 500; cursor: pointer; transition: all 0.2s;
}
.btn-primary { background: var(--primary); color: white; }
.btn-primary:hover { background: var(--primary-hover); }
.btn-primary:disabled { background: #94a3b8; cursor: not-allowed; }
.btn-outline { background: white; color: var(--primary); border: 1.5px solid var(--primary); }
.btn-outline:hover { background: #eff6ff; }
.btn-success { background: var(--success); color: white; }
.btn-success:hover { background: #15803d; }
.btn-group { display: flex; gap: 12px; margin-top: 16px; flex-wrap: wrap; }

.ws-search {
    width: 100%; padding: 10px 14px; border: 1.5px solid var(--border);
    border-radius: 6px; font-size: 0.95rem; margin-bottom: 12px;
}
.ws-search:focus { outline: none; border-color: var(--primary); box-shadow: 0 0 0 3px rgba(37,99,235,0.1); }
.ws-list { display: flex; flex-wrap: wrap; gap: 8px; max-height: 200px; overflow-y: auto; padding: 4px; }
.ws-chip {
    padding: 8px 16px; border: 1.5px solid var(--border); border-radius: 20px;
    cursor: pointer; font-size: 0.9rem; transition: all 0.2s; user-select: none;
}
.ws-chip:hover { border-color: var(--primary); background: #eff6ff; }
.ws-chip.selected { background: var(--primary); color: white; border-color: var(--primary); }

.mapping-table { width: 100%; border-collapse: collapse; font-size: 0.9rem; }
.mapping-table th, .mapping-table td { padding: 10px 14px; border: 1px solid var(--border); text-align: left; }
.mapping-table th { background: #f8fafc; font-weight: 600; color: var(--text-secondary); font-size: 0.85rem; }
.mapping-table select { width: 100%; padding: 6px 10px; border: 1.5px solid var(--border); border-radius: 4px; font-size: 0.85rem; }
.mapping-table .auto-match { color: var(--success); font-size: 0.8rem; font-weight: 600; }
.mapping-table .no-match { color: var(--warning); font-size: 0.8rem; font-weight: 600; }

.result-box {
    display: none; padding: 20px; background: #f0fdf4; border-radius: var(--radius);
    text-align: center; border: 1.5px solid #bbf7d0;
}
.result-box.show { display: block; }
.result-box .check-icon { font-size: 3rem; color: var(--success); margin-bottom: 8px; }
.result-box h3 { color: var(--success); margin-bottom: 8px; }

.spinner {
    display: none; width: 20px; height: 20px;
    border: 2.5px solid var(--border); border-top-color: var(--primary);
    border-radius: 50%; animation: spin 0.6s linear infinite;
}
@keyframes spin { to { transform: rotate(360deg); } }

.alert { padding: 12px 16px; border-radius: 6px; margin-bottom: 16px; font-size: 0.9rem; }
.alert-error { background: #fef2f2; color: var(--danger); border: 1px solid #fecaca; }
.alert-info { background: #eff6ff; color: var(--primary); border: 1px solid #bfdbfe; }
.hidden { display: none !important; }

@media (max-width: 768px) {
    .steps { flex-direction: column; }
    .container { padding: 12px; }
    .card { padding: 16px; }
}
</style>
</head>
<body>

<div class="header">
    <h1>📋 T.QM.013 检验指导书生成器</h1>
    <div class="subtitle">根据版本控制计划，自动生成工位检验指导书</div>
</div>

<div class="container">

    <div class="steps" id="stepIndicator">
        <div class="step active" data-step="1"><span class="step-num">1</span> 上传控制计划</div>
        <div class="step" data-step="2"><span class="step-num">2</span> 选择工位/OP</div>
        <div class="step" data-step="3"><span class="step-num">3</span> 确认列映射</div>
        <div class="step" data-step="4"><span class="step-num">4</span> 生成 &amp; 下载</div>
    </div>

    <div id="globalAlert" class="hidden"></div>

    <!-- Step 1 -->
    <div class="card" id="step1Card">
        <h2>📁 步骤 1：上传版本控制计划</h2>
        <div class="upload-zone" id="uploadZone">
            <div class="upload-icon">📤</div>
            <p>拖拽 Excel 文件到此处，或 <span class="browse-link" id="browseLink">点击浏览</span></p>
            <p style="font-size:0.8rem;margin-top:6px;">支持 .xlsx / .xlsm / .xls 格式</p>
            <input type="file" id="fileInput" accept=".xlsx,.xlsm,.xls">
        </div>
        <div class="file-info" id="fileInfo">
            ✅ <span id="fileName"></span> — <span id="fileStats"></span>
        </div>
        <div class="btn-group" style="justify-content:flex-end;">
            <span class="spinner" id="uploadSpinner"></span>
        </div>
    </div>

    <!-- Step 2 -->
    <div class="card hidden" id="step2Card">
        <h2>🔍 步骤 2：选择目标工位/OP</h2>
        <input type="text" class="ws-search" id="wsSearch" placeholder="搜索工位/OP（输入关键词过滤）...">
        <div class="ws-list" id="wsList"></div>
        <div style="margin-top:8px;color:var(--text-secondary);font-size:0.85rem;">
            已选：<strong id="selectedWsDisplay" style="color:var(--primary);">未选择</strong>
        </div>
        <div class="btn-group" style="justify-content:space-between;">
            <button class="btn btn-outline" onclick="resetAll()">🔄 重新上传</button>
            <button class="btn btn-primary" id="btnNextToMapping" disabled>下一步：确认映射 →</button>
        </div>
    </div>

    <!-- Step 3 -->
    <div class="card hidden" id="step3Card">
        <h2>🔗 步骤 3：确认列映射关系</h2>
        <p style="color:var(--text-secondary);font-size:0.85rem;margin-bottom:12px;">
            系统已自动匹配，如需调整请手动选择。
        </p>
        <div style="overflow-x:auto;">
            <table class="mapping-table">
                <thead><tr><th>T.QM.013 模板列</th><th>控制计划对应列</th><th>状态</th></tr></thead>
                <tbody id="mappingTableBody"></tbody>
            </table>
        </div>
        <div class="btn-group" style="justify-content:space-between;">
            <button class="btn btn-outline" id="btnBackToStep2">← 返回选择工位</button>
            <button class="btn btn-primary" id="btnGenerate">
                🚀 生成检验指导书 <span class="spinner" id="generateSpinner"></span>
            </button>
        </div>
    </div>

    <!-- Step 4 -->
    <div class="card hidden" id="step4Card">
        <h2>✅ 步骤 4：生成结果</h2>
        <div class="result-box" id="resultBox">
            <div class="check-icon">✅</div>
            <h3>检验指导书生成成功！</h3>
            <p id="resultFilename"></p>
        </div>
        <div class="btn-group" style="justify-content:center;">
            <button class="btn btn-success" id="btnDownload">📥 下载检验指导书</button>
            <button class="btn btn-outline" onclick="resetAll()">🔄 生成新的指导书</button>
        </div>
    </div>

</div>

<script>
const $ = s => document.querySelector(s);
const $$ = s => document.querySelectorAll(s);

const STATE = {
    sessionId: null, workstations: [], selectedWorkstation: null,
    allCpHeaders: {}, currentMapping: {}, downloadUrl: null, currentStep: 1
};

function setStep(step) {
    STATE.currentStep = step;
    $$('.step').forEach((el, i) => {
        el.classList.remove('active', 'done');
        if (i+1 < step) el.classList.add('done');
        if (i+1 === step) el.classList.add('active');
    });
    ['step1Card','step2Card','step3Card','step4Card'].forEach((id, i) => {
        document.getElementById(id).classList.toggle('hidden', i+1 !== step);
    });
}

function showAlert(msg, type='error') {
    const el = $('#globalAlert');
    el.className = 'alert alert-' + type;
    el.textContent = msg;
    el.classList.remove('hidden');
    setTimeout(() => el.classList.add('hidden'), 8000);
}

// ===== Step 1: 上传 =====
const uploadZone = $('#uploadZone');
const fileInput = $('#fileInput');
const uploadSpinner = $('#uploadSpinner');

$('#browseLink').addEventListener('click', () => fileInput.click());
uploadZone.addEventListener('click', e => { if (e.target !== $('#browseLink')) fileInput.click(); });
uploadZone.addEventListener('dragover', e => { e.preventDefault(); uploadZone.classList.add('drag-over'); });
uploadZone.addEventListener('dragleave', () => uploadZone.classList.remove('drag-over'));
uploadZone.addEventListener('drop', e => {
    e.preventDefault();
    uploadZone.classList.remove('drag-over');
    if (e.dataTransfer.files.length > 0) handleFile(e.dataTransfer.files[0]);
});
fileInput.addEventListener('change', () => { if (fileInput.files.length > 0) handleFile(fileInput.files[0]); });

async function handleFile(file) {
    const formData = new FormData();
    formData.append('file', file);
    uploadSpinner.style.display = 'inline-block';

    try {
        const resp = await fetch('/api/upload_control_plan', { method: 'POST', body: formData });
        const data = await resp.json();

        if (!data.success) {
            showAlert(data.error || '上传失败');
            return;
        }

        STATE.sessionId = data.session_id;
        STATE.workstations = data.workstations;
        STATE.allCpHeaders = data.headers;

        $('#fileName').textContent = file.name;
        $('#fileStats').textContent = '共 ' + data.total_rows + ' 行数据，' + data.workstations.length + ' 个工位/OP';
        $('#fileInfo').classList.add('show');

        renderWorkstations();
        setStep(2);
    } catch (err) {
        showAlert('网络错误: ' + err.message + '。请确认 Flask 服务已启动，且通过 http://localhost:5000 访问。');
    } finally {
        uploadSpinner.style.display = 'none';
    }
}

// ===== Step 2: 工位选择 =====
function renderWorkstations(filter) {
    filter = (filter || '').toLowerCase();
    const filtered = STATE.workstations.filter(ws => String(ws).toLowerCase().includes(filter));
    const wsList = $('#wsList');
    wsList.innerHTML = filtered.map(ws => {
        const sel = ws === STATE.selectedWorkstation ? ' selected' : '';
        return '<span class="ws-chip' + sel + '" data-ws="' + ws + '">' + ws + '</span>';
    }).join('');
    wsList.querySelectorAll('.ws-chip').forEach(chip => {
        chip.addEventListener('click', () => {
            STATE.selectedWorkstation = chip.dataset.ws;
            $('#selectedWsDisplay').textContent = STATE.selectedWorkstation;
            $('#btnNextToMapping').disabled = false;
            renderWorkstations($('#wsSearch').value);
        });
    });
}

$('#wsSearch').addEventListener('input', e => renderWorkstations(e.target.value));

$('#btnNextToMapping').addEventListener('click', async () => {
    if (!STATE.selectedWorkstation) return;
    await loadColumnMapping();
    setStep(3);
});

// ===== Step 3: 列映射 =====
async function loadColumnMapping() {
    try {
        const resp = await fetch('/api/get_column_mapping', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ session_id: STATE.sessionId })
        });
        const data = await resp.json();
        if (!data.success) { showAlert(data.error); return; }

        STATE.currentMapping = data.mapping;
        const display = data.mapping_display;
        const unmapped = data.unmapped || [];
        const allHeaders = data.all_cp_headers || {};

        let rows = '';
        for (const [tqmCol, info] of Object.entries(display)) {
            rows += '<tr><td><strong>' + info.tqm_label + '</strong></td><td><select data-tqm="' + tqmCol + '" class="mapping-select">';
            rows += '<option value="">-- 不映射 --</option>';
            for (const [col, header] of Object.entries(allHeaders)) {
                const sel = col === info.cp_column ? ' selected' : '';
                rows += '<option value="' + col + '"' + sel + '>' + col + ' — ' + (header || '(空)') + '</option>';
            }
            rows += '</select></td><td><span class="auto-match">✅ 已匹配</span></td></tr>';
        }
        for (const item of unmapped) {
            rows += '<tr><td><strong>' + item.tqm_label + '</strong></td><td><select data-tqm="' + item.tqm_col + '" class="mapping-select">';
            rows += '<option value="">-- 不映射 --</option>';
            for (const [col, header] of Object.entries(allHeaders)) {
                rows += '<option value="' + col + '">' + col + ' — ' + (header || '(空)') + '</option>';
            }
            rows += '</select></td><td><span class="no-match">⚠️ 未匹配</span></td></tr>';
        }
        $('#mappingTableBody').innerHTML = rows;

        $$('.mapping-select').forEach(sel => {
            sel.addEventListener('change', () => {
                STATE.currentMapping = {};
                $$('.mapping-select').forEach(s => {
                    if (s.value) STATE.currentMapping[s.dataset.tqm] = s.value;
                });
            });
        });
    } catch (err) {
        showAlert('网络错误: ' + err.message);
    }
}

$('#btnBackToStep2').addEventListener('click', () => setStep(2));

$('#btnGenerate').addEventListener('click', async () => {
    const spinner = $('#generateSpinner');
    const btn = $('#btnGenerate');
    spinner.style.display = 'inline-block';
    btn.disabled = true;

    try {
        const resp = await fetch('/api/generate', {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({
                session_id: STATE.sessionId,
                workstation: STATE.selectedWorkstation,
                mapping: STATE.currentMapping
            })
        });
        const data = await resp.json();
        if (!data.success) { showAlert(data.error); return; }

        STATE.downloadUrl = data.download_url;
        $('#resultFilename').textContent = '文件名：' + data.filename;
        $('#resultBox').classList.add('show');
        setStep(4);
    } catch (err) {
        showAlert('网络错误: ' + err.message);
    } finally {
        spinner.style.display = 'none';
        btn.disabled = false;
    }
});

// ===== Step 4: 下载 =====
$('#btnDownload').addEventListener('click', () => {
    if (STATE.downloadUrl) window.open(STATE.downloadUrl, '_blank');
});

function resetAll() {
    STATE.sessionId = null; STATE.workstations = []; STATE.selectedWorkstation = null;
    STATE.currentMapping = {}; STATE.downloadUrl = null;
    $('#fileInfo').classList.remove('show');
    $('#wsList').innerHTML = '';
    $('#selectedWsDisplay').textContent = '未选择';
    $('#btnNextToMapping').disabled = true;
    $('#mappingTableBody').innerHTML = '';
    $('#resultBox').classList.remove('show');
    $('#wsSearch').value = '';
    fileInput.value = '';
    setStep(1);
    fetch('/api/cleanup', { method: 'POST' }).catch(() => {});
}

setStep(1);
</script>
</body>
</html>'''


if __name__ == '__main__':
    print("=" * 60)
    print("  T.QM.013 检验指导书生成器")
    print("=" * 60)
    print(f"  模板文件: {os.path.abspath(TEMPLATE_FILE)}")
    if not os.path.exists(TEMPLATE_FILE):
        print(f"  ⚠️  警告: 模板文件未找到！")
        print(f"  请将 T.QM.013 空白模板放在上述路径")
    else:
        print(f"  ✅ 模板文件已找到")
    print(f"  访问地址: http://localhost:5000")
    print("=" * 60)
    app.run(debug=True, host='0.0.0.0', port=5000)
